import os
import openpyxl
import re
import datetime
from file_manager_v101 import get_file_name

# ---- SAMSUNG MODEL DICTIONARY ---- #
SAMSUNG = {
    "A13": "SM-A135F/DS",
    "A22": "SM-A225F",
    "A225": "SM-A225F",
    "A235": "SM-A235F",
    "A32": "SM-A325F",
    "A032": "SM-A032F",
    "A035": "SM-A035F",
    "A037": "SM-A037F/DS",
    "A336": "SM-A336B/DSN",
    "A33": "SM-A336B/DSN",
    "A52": "SM-A525F",
    "A52S": "SM-A528B",
    "A72": "SM-A725F",
    "A12": "SM-A125F",
    "A127": "SM-A127F/DS",
    "A01": "SM-A013G",
    "A11": "SM-A115F",
    "A135": "SM-A135F/DS",
    "FOLD 3": "SM-F926B/DS",
    "NOTE 20": "SM-N980F",
    "NOTE 20 5G": "SM-N981B",
    "NOTE 20 ULTRA": "SM-N986B",
    "G991": "SM-G991B",
    "S21": "SM-G991B",
    "S21+": "SM-G996B",
    "S21 ULTRA": "SM-G998B",
    "E1207": "GT-E1207Y",
    "A022": "SM-A022F/DS",
    "A013 CORE": "SM-A013G/DS",
    "A42": "SM-A426B",
    "A02S": "SM-A025",
    "A025": "SM-A025",
    "A51": "SM-A515F",
    "A53": "SM-A530F/DS",
    "A73": "SM-A730F/DS",
    "F926": "SM-F926B/DS",
    "S20 ULTRA": "SM-G988B",
    "S20 PLUS": "SM-000",
    "S20 FE": "SM-G780G",
    "M12": "SM-M127F",
    "M22": "SM-M225FV/DS",
    "M52": "SM-M526BR",
    "A226": "SM-A226B",
    "A21S": "SM-A217F/DS",
    "A10S": "SM-A107F",
    "NOTE 10": "SM-N970F",
    "NOTE 10 +": "SM-N975F/DS",
    "S901": "SM-S901E/DS",
    "S22": "SM-S901E/DS",
    "S906": "SM-S906E/DS",
    "S22+": "SM-S906E/DS",
    "S908": "SM-S908E/DS",
    "S22 ULTRA": "SM-S908E/DS",
}
TABLETS = {
    "TAB 7": "SM-T736",
    "P615": "SM-P615",
    "T295": "SM-T295",
    "T505": "SM-T505",
    "A7 LITE": "SM-T225",
    "T975": "SM-T975",
    "T870": "SM-T870",
    "T875": "SM-T875",
    "T735": "SM-T735",
    "X205": "SM-X205",
}

# ---- CHOOSING FILES EXCEL FILE --- #
excel_file = get_file_name(file_extension=".xlsx")
wb = openpyxl.load_workbook(excel_file)

# ---- CHOOSING THE SHEET TO WORK ON ---- #
print("here are the sheets(copy and paste them to the input below):")
for sheets in wb:
    print(sheets.title)
sheet_name = input("please enter sheet name: ")
ws = wb[sheet_name]

# ---- PICKING UP COLUMN HEADERS ---- #
column_headers_dictionary = {}
for columns in ws.iter_cols():
    if columns[0].value:
        column_headers = columns[0].value.upper()
        if re.search(r"\b1\b", column_headers):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        elif column_headers.endswith("1"):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        elif column_headers.endswith("2"):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        elif re.search(r"\b2\b", column_headers):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        else:
            for tablets in TABLETS:
                if tablets in column_headers:
                    column_headers_dictionary[column_headers] = columns[0].coordinate

# ---- MATCHING COLUMN HEADERS WITH MODELS ---- #
samsung_phones = list(SAMSUNG.keys())
samsung_tablets = list(TABLETS.keys())
dict_list = samsung_phones + samsung_tablets
matched_phone_dict = {}
matched_tablet_dict = {}
for phones in dict_list:
    for phones2 in column_headers_dictionary:
        if phones.replace(" ", "") in phones2.replace(" ", ""):
            try:
                matched_phone_dict[phones2] = [SAMSUNG[phones], column_headers_dictionary[phones2]]
            except KeyError:
                matched_tablet_dict[phones2] = [TABLETS[phones], column_headers_dictionary[phones2]]

# ---- WRITING THE APPLICATION FORM ---- #
new_wb = openpyxl.Workbook()
new_ws = new_wb.active
IMEI1_count = 0
IMEI2_count = 0
for columns in ws.iter_cols():
    if columns[0].value:
        column_headers = columns[0].value.upper().strip()
        for keys in matched_phone_dict:
            if column_headers.endswith("1"):
                if column_headers == keys.strip() and columns[0].coordinate == matched_phone_dict[keys][1]:
                    for cells in columns[1::]:
                        if cells.value is None:
                            pass
                        elif isinstance(cells.value, int):
                            IMEI1_count += 1
                            new_ws[f"A{IMEI1_count}"] = cells.value
                            new_ws[f"C{IMEI1_count}"] = "SAMSUNG"
                            new_ws[f"D{IMEI1_count}"] = matched_phone_dict[keys][0]
            elif column_headers.endswith("2"):
                if column_headers.replace(" ", "") == keys.replace(" ", "") and\
                        columns[0].coordinate == matched_phone_dict[keys][1]:
                    for cells in columns[1::]:
                        if cells.value and isinstance(cells.value, int):
                            IMEI2_count += 1
                            new_ws[f"B{IMEI2_count}"] = cells.value
                            new_ws[f"C{IMEI2_count}"] = "SAMSUNG"
                            new_ws[f"D{IMEI2_count}"] = matched_phone_dict[keys][0]
print(IMEI1_count)
print(IMEI2_count)

# ---- WRITING IMEI COUNT FOR PHONES ---- #
now = datetime.datetime.now().strftime("%d.%m.%y")
hour = datetime.datetime.now().strftime("%X")
try:
    with open(f"{now}.txt", "a") as file:
        file.write(f"Samsung IMEI 1 count:{IMEI1_count} {hour}\n"
                   f"Samsung IMEI 2 count: {IMEI2_count} {hour}\n")
except FileNotFoundError:
    with open(f"{now}.txt", "w") as file:
        file.write(f"Samsung IMEI 1 count:{IMEI1_count} {hour}\n"
                   f"Samsung IMEI 2 count: {IMEI2_count} {hour}\n")

tablet_IMEI_count = 0
for columns in ws.iter_cols():
    if columns[0].value:
        column_headers = columns[0].value.upper().strip()
        for keys in matched_tablet_dict:
            if column_headers.replace(" ", "") == keys.replace(" ", "") and columns[0].coordinate == matched_tablet_dict[keys][1]:
                for cells in columns[1::]:
                    if cells.value and isinstance(cells.value, int):
                        IMEI1_count += 1
                        tablet_IMEI_count += 1
                        new_ws[f"A{IMEI1_count}"] = cells.value
                        new_ws[f"C{IMEI1_count}"] = "SAMSUNG"
                        new_ws[f"D{IMEI1_count}"] = matched_tablet_dict[keys][0]
print(tablet_IMEI_count)

# ---- WRITING IMEI COUNT FOR TABLETS ---- #
now = datetime.datetime.now().strftime("%d.%m.%y")
try:
    with open(f"{now}.txt", "a") as file:
        file.write(f"Samsung Tablets:{tablet_IMEI_count} {hour}\n")
except FileNotFoundError:
    with open(f"{now}.txt", "w") as file:
        file.write(f"Samsung Tablets:{tablet_IMEI_count} {hour}\n")

new_wb.save(f"Samsung {now} .xlsx")
