#TODO: Create APPLE, SAMSUNG, XIAOMI CLASSES AND HAVE THIS PROJECT COPY EVERYTHING TO AN EXCEL FILE
import openpyxl
import re
from file_manager_v101 import get_file_name, new_file_name

APPLE = {
    "11": "A2221",
    "12": "A2403",
    "12 PRO": "A2407",
    "12 PRO MAX": "A2411",
    "13": "A2631",
    "13 PRO": "A2636",
    "13 PRO MAX": "A2643"
}

XIAOMI = {
    "X3": "M2007J20CG",
    "X3 GT": "21061110AG",
    "X3 PRO": "M2102J20SG",
    "M3": "M2010J19CG",
    "M3 PRO": "M2103K19PG",
    "M4 PRO": "21091116AG",
    "F3": "M2012K11AG",
    "REDMI 9": "M2004J19G",
    "REDMI 10": "M2101K7AG",
    "NOTE 9": "M2003J15SS",
    "NOTE 9 NFC": "M2003J15SG",
    "NOTE 8": "M1908C3JGG",
    "NOTE 10": "M2101K7AG",
    "NOTE 10S": "M2101K7BG",
    "NOTE 10 PRO": "M2101K6G",
    "NOTE 11": "2201117TG",
    "NOTE 11S": "2201117SG",
    "MI 11": "M2011K2G",
    "MI 11 LITE": "M2101K9AG",
    "10T": "M2007J3SY",
    "10T PRO": "M2007J3SG",
    "10T LITE": "M2007J17G",
    "10 LITE": "M2002J9G",
    "11 LITE": "M2101K9AG",
    "11T": "21081111RG",
    "11T PRO": "2107113SG",
    "9A": "M2006C3LG",
    "9C": "M2006C3MG",
    "9T": "M2010J19SG",
    "NOTE 9T": "M2007J22G",
    "MI 12": "2201123G",
    "MI 12 PRO": "2201122G",
    "MI 12X": "2112123AG",
    "NOTE 11 PRO": "2201116TG",
    "NOTE 11 PRO PLUS": "21091116UG",
    "POCO X4 PRO": "2201116PG",
    "10C": "220333QAG",


}

def apple(imei_ws, bthk_ws, row_no):
    # ---- PICKING UP COLUMN HEADERS ---- #
    column_headers_dictionary = {}
    for columns in imei_ws.iter_cols():
        if columns[0].value:
            column_headers = columns[0].value.upper()
            if re.search(r"\b1\b", column_headers):
                column_headers_dictionary[column_headers] = columns[0].coordinate
            elif column_headers.endswith("1"):
                column_headers_dictionary[column_headers] = columns[0].coordinate
            elif column_headers.endswith("2"):
                column_headers_dictionary[column_headers] = columns[0].coordinate
            elif re.search(r"\b2\b", column_headers):
                column_headers_dictionary[column_headers] = columns[0].coordinate
    # ---- MATCHING COLUMN HEADERS WITH MODELS ---- #
    dict_list = list(APPLE.keys())
    matched_dict = {}
    for phones in dict_list:
        for col_heads in column_headers_dictionary:
            if re.search(rf"{phones}\s*\d+", col_heads):
                matched_dict[col_heads] = [APPLE[phones], column_headers_dictionary[col_heads]]

                # ---- WRITING THE APPLICATION FORM ---- #
    IMEI1_count = row_no
    IMEI2_count = row_no

    for columns in imei_ws.iter_cols():
        if columns[0].value:
            column_headers = columns[0].value.upper()
            for keys in matched_dict:
                if column_headers.endswith("1"):
                    if column_headers == keys and columns[0].coordinate == matched_dict[keys][1]:
                        for cells in columns[1::]:
                            if isinstance(cells.value, int):
                                IMEI1_count += 1
                                bthk_ws[f"a{IMEI1_count}"] = cells.value
                                bthk_ws[f"C{IMEI1_count}"] = "APPLE"
                                bthk_ws[f"D{IMEI1_count}"] = matched_dict[keys][0]
                elif column_headers.endswith("2"):
                    if column_headers.replace(" ", "") == keys.replace(" ", "") and \
                            columns[0].coordinate == matched_dict[keys][1]:
                        for cells in columns[1::]:
                            if isinstance(cells.value, int):
                                IMEI2_count += 1
                                bthk_ws[f"B{IMEI2_count}"] = cells.value
                                bthk_ws[f"C{IMEI2_count}"] = "APPLE"
                                bthk_ws[f"D{IMEI2_count}"] = matched_dict[keys][0]
    print(f"The number of first IMEI: {IMEI1_count}\n"
          f"The number of second IMEI: {IMEI2_count}")
    return bthk_ws.max_row


def xiaomi(imei_ws,bthk_ws,row_no):
    # ---- XIAOMI MODEL DICTIONARY ---- #

    # ---- PICKING UP COLUMN HEADERS ---- #
    column_headers_dictionary = {}
    for columns in imei_ws.iter_cols():
        if columns[0].value is None:
            pass
        else:
            column_headers = columns[0].value.upper()
            if re.search(r"\b1\b", column_headers):
                column_headers_dictionary[column_headers] = columns[0].coordinate
            elif column_headers.endswith("1"):
                column_headers_dictionary[column_headers] = columns[0].coordinate
            elif column_headers.endswith("2"):
                column_headers_dictionary[column_headers] = columns[0].coordinate
            elif re.search(r"\b2\b", column_headers):
                column_headers_dictionary[column_headers] = columns[0].coordinate

    # ---- MATCHING COLUMN HEADERS WITH MODELS ---- #
    dict_list = list(XIAOMI.keys())
    matched_dict = {}
    for phones in dict_list:
        for phones2 in column_headers_dictionary:
            if phones.replace(" ", "") in phones2.replace(" ", ""):
                matched_dict[phones2] = [XIAOMI[phones], column_headers_dictionary[phones2]]

    # ---- WRITING THE APPLICATION FORM ---- #
    IMEI1_count = row_no
    IMEI2_count = row_no
    for columns in imei_ws.iter_cols():
        if columns[0].value:
            column_headers = columns[0].value.upper().strip()
            for keys in matched_dict:
                if column_headers.endswith("1"):
                    if column_headers == keys.strip() and columns[0].coordinate == matched_dict[keys][1]:
                        for cells in columns[1::]:
                            if cells.value is None:
                                pass
                            elif isinstance(cells.value, int):
                                IMEI1_count += 1
                                bthk_ws[f"A{IMEI1_count}"] = cells.value
                                bthk_ws[f"C{IMEI1_count}"] = "XIAOMI"
                                bthk_ws[f"D{IMEI1_count}"] = matched_dict[keys][0]
                elif column_headers.endswith("2"):
                    if column_headers.replace(" ", "") == keys.replace(" ", "") and columns[0].coordinate == \
                            matched_dict[keys][1]:
                        for cells in columns[1::]:
                            if isinstance(cells.value, int):
                                IMEI2_count += 1
                                bthk_ws[f"B{IMEI2_count}"] = cells.value
                                bthk_ws[f"C{IMEI2_count}"] = "XIAOMI"
                                bthk_ws[f"D{IMEI2_count}"] = matched_dict[keys][0]

    print(f"The number of first IMEI: {IMEI1_count}\n"
          f"The number of second IMEI: {IMEI2_count}")
    return bthk_ws.max_row


def imei_writer(imei_ws, bthk_ws, row_no, model_dict, brand_name):
    # ---- PICKING UP COLUMN HEADERS ---- #
    column_headers_dictionary = {}
    for columns in imei_ws.iter_cols():
        if columns[0].value:
            column_headers = columns[0].value.upper()
            if "+" in column_headers:
                new_column_headers = column_headers.split(" ")
                index_of_plus = new_column_headers.index("+")
                new_column_headers.remove("+")
                new_column_headers.insert(index_of_plus, "PLUS")
                column_headers = " ".join(new_column_headers)
                column_headers_dictionary[column_headers] = columns[0].coordinate
            else:
                column_headers_dictionary[column_headers] = columns[0].coordinate
        # if re.search(r"\b1\b", column_headers):
        #     column_headers_dictionary[column_headers] = columns[0].coordinate
        # elif column_headers.endswith("1"):
        #     column_headers_dictionary[column_headers] = columns[0].coordinate
        # elif column_headers.endswith("2"):
        #     column_headers_dictionary[column_headers] = columns[0].coordinate
        # elif re.search(r"\b2\b", column_headers):
        #         column_headers_dictionary[column_headers] = columns[0].coordinate

    # ---- MATCHING COLUMN HEADERS WITH MODELS ---- #
    dict_list = list(model_dict.keys())
    matched_dict = {}
    for phones in dict_list:
        for col_heads in column_headers_dictionary:
            if re.search(rf"{phones}\s*\d+", col_heads):
                matched_dict[col_heads] = [model_dict[phones], column_headers_dictionary[col_heads]]

    # ---- WRITING THE APPLICATION FORM ---- #
    IMEI1_count = row_no
    IMEI2_count = row_no
    for columns in imei_ws.iter_cols():
        if columns[0].value:
            column_headers = columns[0].value.upper().strip()
            for keys in matched_dict:
                if column_headers.endswith("1"):
                    if columns[0].coordinate == matched_dict[keys][1]:
                        for cells in columns[1::]:
                            if isinstance(cells.value, int):
                                IMEI1_count += 1
                                bthk_ws[f"A{IMEI1_count}"] = cells.value
                                bthk_ws[f"C{IMEI1_count}"] = brand_name.upper()
                                bthk_ws[f"D{IMEI1_count}"] = matched_dict[keys][0]
                elif column_headers.endswith("2"):
                    if columns[0].coordinate == matched_dict[keys][1]:
                        for cells in columns[1::]:
                            if isinstance(cells.value, int):
                                IMEI2_count += 1
                                bthk_ws[f"B{IMEI2_count}"] = cells.value
                                bthk_ws[f"C{IMEI2_count}"] = brand_name.upper()
                                bthk_ws[f"D{IMEI2_count}"] = matched_dict[keys][0]

    print(f"The number of first IMEI: {IMEI1_count}\n"
          f"The number of second IMEI: {IMEI2_count}")
    return bthk_ws.max_row


excel_file = get_file_name(file_extension=".xlsx")
wb = openpyxl.load_workbook(excel_file)
imei_ws = wb
new_wb = openpyxl.Workbook()
new_ws = new_wb.active
ws_list = [sheets for sheets in imei_ws]
for num, sheets in enumerate(imei_ws, 1):
    print(num, sheets.title)
run_machine = True
row_no = 0
while run_machine:
    question_1 = input("Please enter the number of sheet you want to work on and"
                       " one of the following brand names (apple,samsung,xiaomi): ")
    sheet_number = 0
    model_name = ""
    for var in question_1.split(" "):
        if var.isdigit():
            sheet_number = int(var)
        else:
            model_name = var
    if model_name == "apple":
        row_no = imei_writer(wb[ws_list[sheet_number - 1].title], new_ws, row_no, APPLE, model_name)
    if model_name == "xiaomi":
        row_no = imei_writer(wb[ws_list[sheet_number - 1].title], new_ws, row_no, XIAOMI, model_name)
    question_2 = input("Do you want to continue y/n? ")
    if question_2.lower() == "y":
        continue
    else:
        run_machine = False
new_file_name = new_file_name("xlsx")
new_wb.save(new_file_name)

