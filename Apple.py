import openpyxl
import re
import os
import datetime
from file_manager_v101 import get_file_name
from collections import defaultdict


#TODO.2: MAKE ALL OF THESE FUNCTIONS

# ---- APPLE MODEL DICTIONARY ---- #
APPLE = {
    "11": "A2221",
    "12": "A2403",
    "12 PRO": "A2407",
    "12 PRO MAX": "A2411",
    "13": "A2631",
    "13 PRO": "A2636",
    "13 PRO MAX": "A2643"
        }

# ---- CHOOSING FILES EXCEL FILE --- #

excel_file = get_file_name(file_extension=".xlsx")
wb = openpyxl.load_workbook(excel_file)
# ---- CHOOSING THE SHEET TO WORK ON ---- #
print("here are the sheets(copy and paste them to the input below):")
for sheets in wb:
    print(sheets.title)
sheet_name = input("please enter sheet name: ")
ws = wb[sheet_name]


# ---- PICKING UP COLUMN HEADERS ---- #
column_headers_dictionary = {}
for columns in ws.iter_cols():
    if columns[0].value:
        column_headers = columns[0].value.upper()
        if re.search(r"\b1\b", column_headers):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        elif column_headers.endswith("1"):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        elif column_headers.endswith("2"):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        elif re.search(r"\b2\b", column_headers):
            column_headers_dictionary[column_headers] = columns[0].coordinate

# ---- MATCHING COLUMN HEADERS WITH MODELS ---- #
dict_list = list(APPLE.keys())
matched_dict = {}
for phones in dict_list:
    for col_heads in column_headers_dictionary:
        if re.search(rf"{phones}\s*\d+", col_heads):
            matched_dict[col_heads] = [APPLE[phones], column_headers_dictionary[col_heads]]

# ---- WRITING THE APPLICATION FORM ---- #
IMEI1_count = 0
IMEI2_count = 0
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

for columns in ws.iter_cols():
    if columns[0].value is None:
        pass
    else:
        column_headers = columns[0].value.upper()
        for keys in matched_dict:
            if column_headers.endswith("1"):
                if column_headers == keys and columns[0].coordinate == matched_dict[keys][1]:
                    for cells in columns[1::]:
                        if cells.value is None:
                            pass
                        elif isinstance(cells.value, int):
                            IMEI1_count += 1
                            new_ws[f"a{IMEI1_count}"] = cells.value
                            new_ws[f"C{IMEI1_count}"] = "APPLE"
                            new_ws[f"D{IMEI1_count}"] = matched_dict[keys][0]
            elif column_headers.endswith("2"):
                if column_headers.replace(" ", "") == keys.replace(" ", "") and\
                        columns[0].coordinate == matched_dict[keys][1]:
                    for cells in columns[1::]:
                        if cells.value is None:
                            pass
                        elif isinstance(cells.value, int):
                            IMEI2_count += 1
                            new_ws[f"B{IMEI2_count}"] = cells.value
                            new_ws[f"C{IMEI2_count}"] = "APPLE"
                            new_ws[f"D{IMEI2_count}"] = matched_dict[keys][0]
print(f"The number of first IMEI: {IMEI1_count}\n"
      f"The number of second IMEI: {IMEI2_count}")

# ---- WRITING IMEI COUNT ---- #
now = datetime.datetime.now().strftime("%d.%m.%y")
hour = datetime.datetime.now().strftime("%X")
try:
    with open(f"{now}.txt", "a") as file:
        file.write(f"Iphone IMEI1 count: {IMEI1_count}\nIphone IMEI2 count: {IMEI2_count} {hour}\n")
except FileNotFoundError:
    with open(f"{now}.txt", "w") as file:
        file.write(f"Iphone IMEI1 count: {IMEI1_count}\nIphone IMEI2 count: {IMEI2_count} {hour}\n")
new_wb.save(f"Apple {now} .xlsx")

