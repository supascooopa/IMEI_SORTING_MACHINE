import os
import openpyxl
import re
import datetime
from file_manager_v101 import get_file_name
# ---- XIAOMI MODEL DICTIONARY ---- #
XIAOMI = {
    "X3": "M2007J20CG",
    "X3 GT": "21061110AG",
    "X3 PRO": "M2102J20SG",
    "M3": "M2010J19CG",
    "M3 PRO": "M2103K19PG",
    "M4 PRO": "21091116AG",
    "F3": "M2012K11AG",
    "REDMI 9": "M2004J19G",
    "REDMI 10": "M2101K7AG",
    "NOTE 9": "M2003J15SS",
    "NOTE 9 NFC": "M2003J15SG",
    "NOTE 8": "M1908C3JGG",
    "NOTE 10": "M2101K7AG",
    "NOTE 10S": "M2101K7BG",
    "NOTE 10 PRO": "M2101K6G",
    "NOTE 11": "2201117TG",
    "MI 11": "M2011K2G",
    "MI 11 LITE": "M2101K9AG",
    "10T": "M2007J3SY",
    "10T PRO": "M2007J3SG",
    "10T LITE": "M2007J17G",
    "10 LITE": "M2002J9G",
    "11 LITE": "M2101K9AG",
    "11T": "21081111RG",
    "11T PRO": "2107113SG",
    "9A": "M2006C3LG",
    "9C": "M2006C3MG",
    "9T": "M2010J19SG",
    "NOTE 9T": "M2007J22G",
    "MI 12": "2201123G",
    "MI 12 PRO": "2201122G",
    "MI 12X": "2112123AG",
    "NOTE 11 PRO": "2201116TG",
    "NOTE 11 PRO+": "21091116UG",
    "POCO X4 PRO": "2201116PG",
    "10C": "220333QAG",


}
# ---- CHOOSING FILES EXCEL FILE --- #
# print("below are the files in your current directory:")
# file_index = 1
# for files in os.listdir():
#     if files.endswith("xlsx"):
#         print(f"{file_index}. {files}")
#         file_index += 1
excel_file = get_file_name(file_extension=".xlsx")
wb = openpyxl.load_workbook(excel_file)

# ---- CHOOSING THE SHEET TO WORK ON ---- #
print("here are the sheets(copy and paste them to the input below):")
for sheets in wb:
    print(sheets.title)
sheet_name = input("please enter sheet name: ")
ws = wb[sheet_name]

# ---- PICKING UP COLUMN HEADERS ---- #
column_headers_dictionary = {}
for columns in ws.iter_cols():
    if columns[0].value:
        column_headers = columns[0].value.upper()
        if re.search(r"\b1\b", column_headers):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        elif column_headers.endswith("1"):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        elif column_headers.endswith("2"):
            column_headers_dictionary[column_headers] = columns[0].coordinate
        elif re.search(r"\b2\b", column_headers):
            column_headers_dictionary[column_headers] = columns[0].coordinate

# ---- MATCHING COLUMN HEADERS WITH MODELS ---- #
dict_list = list(XIAOMI.keys())
matched_dict = {}
for phones in dict_list:
    for phones2 in column_headers_dictionary:
        if phones.replace(" ", "") in phones2.replace(" ", ""):
            matched_dict[phones2] = [XIAOMI[phones], column_headers_dictionary[phones2]]

# ---- WRITING THE APPLICATION FORM ---- #
new_wb = openpyxl.Workbook()
new_ws = new_wb.active
IMEI1_count = 0
IMEI2_count = 0
for columns in ws.iter_cols():
    if columns[0].value:
        column_headers = columns[0].value.upper().strip()
        for keys in matched_dict:
            if column_headers.endswith("1"):
                if column_headers == keys.strip() and columns[0].coordinate == matched_dict[keys][1]:
                    for cells in columns[1::]:
                        if isinstance(cells.value, int):
                            IMEI1_count += 1
                            new_ws[f"A{IMEI1_count}"] = cells.value
                            new_ws[f"C{IMEI1_count}"] = "XIAOMI"
                            new_ws[f"D{IMEI1_count}"] = matched_dict[keys][0]
            elif column_headers.endswith("2"):
                if column_headers.replace(" ", "") == keys.replace(" ", "") and columns[0].coordinate == matched_dict[keys][1]:
                    for cells in columns[1::]:
                        if isinstance(cells.value, int):
                            IMEI2_count += 1
                            new_ws[f"B{IMEI2_count}"] = cells.value
                            new_ws[f"C{IMEI2_count}"] = "XIAOMI"
                            new_ws[f"D{IMEI2_count}"] = matched_dict[keys][0]
print(f"The number of first IMEI: {IMEI1_count}\n"
      f"The number of second IMEI: {IMEI2_count}")

# ---- WRITING IMEI COUNT ---- #
now = datetime.datetime.now().strftime("%d.%m.%y")
hour = datetime.datetime.now().strftime("%X")
try:
    with open(f"{now}.txt", "a") as file:
        file.write(f"Xiaomi IMEI1 count:{IMEI1_count}\nXiaomi IMEI2 count:{IMEI2_count} {hour}\n")
except FileNotFoundError:
    with open(f"{now}.txt", "w") as file:
        file.write(f"Xiaomi IMEI1 count:{IMEI1_count}\nXiaomi IMEI2 count:{IMEI2_count} {hour}\n")
new_wb.save(f"Xiaomi {now} .xlsx")
